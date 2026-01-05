#!/usr/bin/env python3
"""Pytest suite for function_size_calculator.py."""

import json
import os
import sys
from contextlib import redirect_stderr, redirect_stdout
from io import StringIO
from pathlib import Path

import pytest

# Add parent directory to path to import the module
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from function_size_calculator import (
    ExcelWriter,
    FunctionInfo,
    JavaParser,
    JavaScriptParser,
    JSONWriter,
    PythonParser,
    is_test_file,
    print_progress_bar,
    scan_single_repository,
)

try:
    import openpyxl
except ImportError:  # pragma: no cover - dependency is optional in CI
    openpyxl = None


@pytest.fixture
def fixtures_dir() -> Path:
    """Return path to test fixtures."""
    return Path(__file__).parent / "fixtures"


class TestFunctionInfo:
    """Tests for FunctionInfo class."""

    def test_function_info_creation(self):
        func = FunctionInfo(
            name="testFunction",
            file_path="test.js",
            start_line=1,
            end_line=10,
            size=10,
        )

        assert func.name == "testFunction"
        assert func.file_path == "test.js"
        assert func.start_line == 1
        assert func.end_line == 10
        assert func.size == 10

    def test_function_info_repr(self):
        func = FunctionInfo("myFunc", "file.js", 5, 15, 11)
        repr_str = repr(func)

        assert "myFunc" in repr_str
        assert "file.js" in repr_str
        assert "11" in repr_str

    def test_function_info_to_dict(self):
        func = FunctionInfo("testFunc", "test.py", 10, 20, 11)
        func_dict = func.to_dict()

        assert func_dict["name"] == "testFunc"
        assert func_dict["file_path"] == "test.py"
        assert func_dict["start_line"] == 10
        assert func_dict["end_line"] == 20
        assert func_dict["size"] == 11


class TestJavaScriptParser:
    """Tests for JavaScriptParser."""

    @pytest.fixture(autouse=True)
    def _setup(self, fixtures_dir: Path):
        self.js_file = fixtures_dir / "sample.js"
        self.ts_file = fixtures_dir / "sample.ts"

    def test_parse_javascript_file(self):
        functions = JavaScriptParser.parse_functions(str(self.js_file))

        assert len(functions) > 0

        func_names = [f.name for f in functions]
        assert "simpleFunction" in func_names
        assert "largeFunction" in func_names
        assert "arrowFunction" in func_names
        assert "asyncArrowFunction" in func_names
        assert "outerFunction" in func_names

    def test_parse_typescript_file(self):
        functions = JavaScriptParser.parse_functions(str(self.ts_file))

        assert len(functions) > 0

        func_names = [f.name for f in functions]
        assert "typedFunction" in func_names

    def test_function_size_calculation(self):
        functions = JavaScriptParser.parse_functions(str(self.js_file))

        simple = next((f for f in functions if f.name == "simpleFunction"), None)
        assert simple is not None
        assert simple.size == 3

        large = next((f for f in functions if f.name == "largeFunction"), None)
        assert large is not None
        assert large.size > simple.size

    def test_parse_nonexistent_file(self):
        with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
            functions = JavaScriptParser.parse_functions("/nonexistent/file.js")

        assert len(functions) == 0

    def test_function_line_numbers(self):
        functions = JavaScriptParser.parse_functions(str(self.js_file))

        for func in functions:
            assert func.start_line > 0
            assert func.end_line >= func.start_line
            assert func.size > 0


class TestJavaParser:
    """Tests for JavaParser."""

    @pytest.fixture(autouse=True)
    def _setup(self, fixtures_dir: Path):
        self.java_file = fixtures_dir / "Sample.java"

    def test_parse_java_file(self):
        functions = JavaParser.parse_functions(str(self.java_file))

        assert len(functions) > 0

        func_names = [f.name for f in functions]
        assert "publicMethod" in func_names
        assert "privateMethod" in func_names
        assert "protectedStaticMethod" in func_names
        assert "largeMethod" in func_names

    def test_java_method_modifiers(self):
        functions = JavaParser.parse_functions(str(self.java_file))
        func_names = [f.name for f in functions]

        assert "publicStaticFinalMethod" in func_names
        assert "synchronizedMethod" in func_names
        assert "methodWithException" in func_names

    def test_java_function_size(self):
        functions = JavaParser.parse_functions(str(self.java_file))

        large = next((f for f in functions if f.name == "largeMethod"), None)
        assert large is not None
        assert large.size >= 10

    def test_parse_nonexistent_java_file(self):
        with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
            functions = JavaParser.parse_functions("/nonexistent/Sample.java")

        assert len(functions) == 0


class TestPythonParser:
    """Tests for PythonParser."""

    @pytest.fixture(autouse=True)
    def _setup(self, fixtures_dir: Path):
        self.py_file = fixtures_dir / "sample.py"

    def test_parse_python_file(self):
        functions = PythonParser.parse_functions(str(self.py_file))

        assert len(functions) > 0

        func_names = [f.name for f in functions]
        assert "simple_function" in func_names
        assert "large_function" in func_names
        assert "async_function" in func_names
        assert "typed_function" in func_names
        assert "outer_function" in func_names

    def test_python_class_methods(self):
        functions = PythonParser.parse_functions(str(self.py_file))
        func_names = [f.name for f in functions]

        assert "method_in_class" in func_names
        assert "another_method" in func_names
        assert "async_method" in func_names

    def test_python_function_size(self):
        functions = PythonParser.parse_functions(str(self.py_file))

        simple = next((f for f in functions if f.name == "simple_function"), None)
        assert simple is not None
        assert simple.size == 2

        large = next((f for f in functions if f.name == "large_function"), None)
        assert large is not None
        assert large.size > simple.size

    def test_parse_nonexistent_python_file(self):
        with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
            functions = PythonParser.parse_functions("/nonexistent/sample.py")

        assert len(functions) == 0

    def test_python_function_line_numbers(self):
        functions = PythonParser.parse_functions(str(self.py_file))

        for func in functions:
            assert func.start_line > 0
            assert func.end_line >= func.start_line
            assert func.size > 0


class TestExcelWriter:
    """Tests for ExcelWriter."""

    @pytest.fixture(autouse=True)
    def _setup(self, tmp_path: Path):
        self.output_file = tmp_path / "test_output.xlsx"
        self.sample_functions = [
            FunctionInfo("func1", "file1.js", 1, 10, 10),
            FunctionInfo("func2", "file2.js", 1, 20, 20),
            FunctionInfo("func3", "file3.js", 1, 15, 15),
            FunctionInfo("func4", "file4.js", 1, 5, 5),
            FunctionInfo("func5", "file5.js", 1, 8, 8),
            FunctionInfo("func6", "file6.js", 1, 12, 12),
        ]

    @pytest.mark.skipif(openpyxl is None, reason="openpyxl not available")
    def test_write_results_single_repo(self):
        repo_results = {"test-repo": self.sample_functions}

        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, str(self.output_file))

        assert self.output_file.exists()

        wb = openpyxl.load_workbook(self.output_file)
        assert "test-repo" in wb.sheetnames

        ws = wb["test-repo"]
        assert ws.cell(1, 1).value == "Rank"
        assert ws.cell(1, 2).value == "Function Name"
        assert ws.cell(1, 6).value == "Lines of Code"

        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == "func2"
        assert ws.cell(2, 6).value == 20

        assert ws.cell(6, 1).value == 5
        wb.close()

    @pytest.mark.skipif(openpyxl is None, reason="openpyxl not available")
    def test_write_results_multiple_repos(self):
        repo_results = {
            "repo1": self.sample_functions[:3],
            "repo2": self.sample_functions[3:],
        }

        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, str(self.output_file))

        assert self.output_file.exists()

        wb = openpyxl.load_workbook(self.output_file)
        assert "repo1" in wb.sheetnames
        assert "repo2" in wb.sheetnames
        wb.close()

    @pytest.mark.skipif(openpyxl is None, reason="openpyxl not available")
    def test_sanitize_sheet_name(self):
        repo_results = {
            "very/long/repository/name/that/exceeds/thirty/one/characters": self.sample_functions[:1]
        }

        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, str(self.output_file))

        wb = openpyxl.load_workbook(self.output_file)
        sheet_names = wb.sheetnames
        assert len(sheet_names) == 1
        assert len(sheet_names[0]) <= 31
        assert "/" not in sheet_names[0]
        wb.close()

    @pytest.mark.skipif(openpyxl is None, reason="openpyxl not available")
    def test_top_n_parameter(self):
        repo_results = {"test-repo": self.sample_functions}

        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, str(self.output_file), top_n=3)

        wb = openpyxl.load_workbook(self.output_file)
        ws = wb["test-repo"]

        data_rows = 0
        for row in range(2, 10):
            value = ws.cell(row, 1).value
            if value and isinstance(value, int):
                data_rows += 1

        assert data_rows == 3
        wb.close()

    @pytest.mark.skipif(openpyxl is None, reason="openpyxl not available")
    def test_min_size_filter(self):
        repo_results = {"test-repo": self.sample_functions}

        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, str(self.output_file), top_n=10, min_size=10)

        wb = openpyxl.load_workbook(self.output_file)
        ws = wb["test-repo"]

        data_rows = 0
        for row in range(2, 10):
            value = ws.cell(row, 1).value
            if value and isinstance(value, int):
                data_rows += 1

        assert data_rows == 4

        for row in range(2, 6):
            size = ws.cell(row, 6).value
            if size is not None and isinstance(size, int):
                assert size >= 10

        wb.close()

    @pytest.mark.skipif(openpyxl is None, reason="openpyxl not available")
    def test_summary_statistics(self):
        repo_results = {"test-repo": self.sample_functions}

        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, str(self.output_file))

        wb = openpyxl.load_workbook(self.output_file)
        ws = wb["test-repo"]

        found_summary = any(ws.cell(row, 1).value == "Summary Statistics" for row in range(1, 20))
        assert found_summary, "Summary statistics section not found"
        wb.close()


class TestJSONWriter:
    """Tests for JSONWriter."""

    @pytest.fixture(autouse=True)
    def _setup(self, tmp_path: Path):
        self.output_file = tmp_path / "test_output.json"
        self.sample_functions = [
            FunctionInfo("func1", "file1.js", 1, 10, 10),
            FunctionInfo("func2", "file2.js", 1, 20, 20),
            FunctionInfo("func3", "file3.js", 1, 15, 15),
            FunctionInfo("func4", "file4.js", 1, 5, 5),
            FunctionInfo("func5", "file5.js", 1, 8, 8),
        ]

    def test_write_results_single_repo(self):
        repo_results = {"test-repo": self.sample_functions}

        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, str(self.output_file))

        assert self.output_file.exists()

        data = json.loads(self.output_file.read_text())

        assert "test-repo" in data
        assert "summary" in data["test-repo"]
        assert "top_functions" in data["test-repo"]

        summary = data["test-repo"]["summary"]
        assert summary["total_functions"] == 5
        assert summary["average_size"] > 0
        assert summary["largest_function_size"] == 20
        assert summary["smallest_function_size"] == 5

        top_funcs = data["test-repo"]["top_functions"]
        assert len(top_funcs) == 5
        assert top_funcs[0]["name"] == "func2"
        assert top_funcs[0]["size"] == 20

    def test_write_results_multiple_repos(self):
        repo_results = {
            "repo1": self.sample_functions[:3],
            "repo2": self.sample_functions[3:],
        }

        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, str(self.output_file))

        assert self.output_file.exists()

        data = json.loads(self.output_file.read_text())
        assert "repo1" in data
        assert "repo2" in data

    def test_top_n_parameter(self):
        repo_results = {"test-repo": self.sample_functions}

        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, str(self.output_file), top_n=3)

        data = json.loads(self.output_file.read_text())

        assert len(data["test-repo"]["top_functions"]) == 3

    def test_min_size_filter(self):
        repo_results = {"test-repo": self.sample_functions}

        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, str(self.output_file), top_n=10, min_size=10)

        data = json.loads(self.output_file.read_text())

        top_funcs = data["test-repo"]["top_functions"]
        assert len(top_funcs) == 3
        assert all(func["size"] >= 10 for func in top_funcs)
        assert data["test-repo"]["summary"]["total_functions"] == 3

    def test_min_size_filter_multiple_repos(self):
        repo1_functions = [
            FunctionInfo("large_func", "file1.js", 1, 50, 50),
            FunctionInfo("medium_func", "file2.js", 1, 25, 25),
        ]
        repo2_functions = [
            FunctionInfo("small_func", "file3.js", 1, 5, 5),
            FunctionInfo("tiny_func", "file4.js", 1, 3, 3),
        ]

        repo_results = {"repo1": repo1_functions, "repo2": repo2_functions}

        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, str(self.output_file), top_n=10, min_size=2)

        data = json.loads(self.output_file.read_text())

        assert data["repo1"]["summary"]["total_functions"] == 2
        assert data["repo2"]["summary"]["total_functions"] == 2
        assert len(data["repo1"]["top_functions"]) == 2
        assert len(data["repo2"]["top_functions"]) == 2


@pytest.fixture
def test_repository(tmp_path: Path) -> Path:
    """Create a temporary repository structure for scanning tests."""
    repo_dir = tmp_path / "test_repo"
    src_dir = repo_dir / "src"
    src_dir.mkdir(parents=True)

    js_content = """
function testFunc() {
    console.log("test");
}
"""
    (src_dir / "test.js").write_text(js_content)

    java_content = """
public class Sample {
    public void testMethod() {
        System.out.println("test");
    }
}
"""
    (src_dir / "Sample.java").write_text(java_content)

    python_content = """
def python_function():
    print("hello")
"""
    (src_dir / "sample.py").write_text(python_content)

    return repo_dir


class TestScanRepository:
    """Tests for scan_single_repository function."""

    def test_scan_local_repository(self, test_repository: Path):
        repo_name, functions = scan_single_repository(str(test_repository))

        assert repo_name is not None
        assert len(functions) > 0

        func_names = [f.name for f in functions]
        assert "testFunc" in func_names
        assert "testMethod" in func_names
        assert "python_function" in func_names

    def test_scan_nonexistent_repository(self):
        with redirect_stdout(StringIO()), redirect_stderr(StringIO()):
            repo_name, functions = scan_single_repository("/nonexistent/repo")

        assert repo_name is None
        assert len(functions) == 0

    def test_relative_paths(self, test_repository: Path):
        _, functions = scan_single_repository(str(test_repository))

        for func in functions:
            assert not func.file_path.startswith("/")
            assert not func.file_path.startswith(str(test_repository))


class TestCommandLineArguments:
    """Tests for command-line argument handling."""

    def test_input_file_parsing(self, tmp_path: Path):
        input_file = tmp_path / "repos.txt"
        input_file.write_text(
            "https://github.com/user/repo1.git\n"
            "# This is a comment\n"
            "\n"
            "/path/to/local/repo\n"
        )

        repositories = []
        with open(input_file, encoding="utf-8") as file_handle:
            for line in file_handle:
                line = line.strip()
                if line and not line.startswith("#"):
                    repositories.append(line)

        assert len(repositories) == 2
        assert "https://github.com/user/repo1.git" in repositories
        assert "/path/to/local/repo" in repositories


class TestProgressBar:
    """Tests for print_progress_bar function."""

    def test_progress_bar_zero_total(self, capsys):
        """Progress bar should handle zero total gracefully."""
        print_progress_bar(0, 0, prefix="Test:", suffix="Done")
        captured = capsys.readouterr()
        # Should return early without printing
        assert captured.out == ""

    def test_progress_bar_partial_progress(self, capsys):
        """Progress bar should show partial progress."""
        print_progress_bar(5, 10, prefix="Progress:", suffix="Done")
        captured = capsys.readouterr()
        assert "50.0%" in captured.out
        assert "Progress:" in captured.out

    def test_progress_bar_complete(self, capsys):
        """Progress bar should print newline when complete."""
        print_progress_bar(10, 10, prefix="Progress:", suffix="Done")
        captured = capsys.readouterr()
        assert "100.0%" in captured.out
        # Should end with newline when complete
        assert captured.out.endswith("\n")


class TestEmptyResults:
    """Tests for handling empty results scenarios."""

    @pytest.fixture(autouse=True)
    def _setup(self, tmp_path: Path):
        self.output_file = tmp_path / "test_output"

    @pytest.mark.skipif(openpyxl is None, reason="openpyxl not available")
    def test_write_empty_repo_excel(self):
        """ExcelWriter should handle empty function list."""
        repo_results = {"empty-repo": []}
        output_file = str(self.output_file) + ".xlsx"

        with redirect_stdout(StringIO()):
            ExcelWriter.write_results(repo_results, output_file)

        assert Path(output_file).exists()
        wb = openpyxl.load_workbook(output_file)
        ws = wb["empty-repo"]
        # Should still have header
        assert ws.cell(1, 1).value == "Rank"
        wb.close()

    def test_write_empty_repo_json(self):
        """JSONWriter should handle empty function list."""
        repo_results = {"empty-repo": []}
        output_file = str(self.output_file) + ".json"

        with redirect_stdout(StringIO()):
            JSONWriter.write_results(repo_results, output_file)

        assert Path(output_file).exists()
        data = json.loads(Path(output_file).read_text())
        assert "empty-repo" in data
        assert data["empty-repo"]["summary"] == {}
        assert data["empty-repo"]["top_functions"] == []


class TestIsTestFile:
    """Tests for is_test_file function."""

    def test_java_test_file_with_test_suffix(self):
        """Should identify Java files ending with Test.java as test files."""
        assert is_test_file(Path("/src/main/java/SampleTest.java")) is True
        assert is_test_file(Path("/src/main/java/UserTest.java")) is True

    def test_java_test_file_with_tests_suffix(self):
        """Should identify Java files ending with Tests.java as test files."""
        assert is_test_file(Path("/src/main/java/SampleTests.java")) is True
        assert is_test_file(Path("/src/main/java/UserTests.java")) is True

    def test_java_non_test_file(self):
        """Should not identify regular Java files as test files."""
        assert is_test_file(Path("/src/main/java/Sample.java")) is False
        assert is_test_file(Path("/src/main/java/User.java")) is False
        # TestUtils.java should NOT be identified as a test file (avoids false positives)
        assert is_test_file(Path("/src/main/java/TestUtils.java")) is False
        assert is_test_file(Path("/src/main/java/TestConstants.java")) is False

    def test_javascript_test_file_with_test_extension(self):
        """Should identify JS/TS files with .test. pattern as test files."""
        assert is_test_file(Path("/src/sample.test.js")) is True
        assert is_test_file(Path("/src/sample.test.ts")) is True
        assert is_test_file(Path("/src/component.test.jsx")) is True
        assert is_test_file(Path("/src/component.test.tsx")) is True

    def test_javascript_spec_file(self):
        """Should identify JS/TS files with .spec. pattern as test files."""
        assert is_test_file(Path("/src/sample.spec.js")) is True
        assert is_test_file(Path("/src/sample.spec.ts")) is True
        assert is_test_file(Path("/src/component.spec.jsx")) is True
        assert is_test_file(Path("/src/component.spec.tsx")) is True

    def test_javascript_non_test_file(self):
        """Should not identify regular JS/TS files as test files."""
        assert is_test_file(Path("/src/sample.js")) is False
        assert is_test_file(Path("/src/sample.ts")) is False
        assert is_test_file(Path("/src/component.jsx")) is False
        assert is_test_file(Path("/src/component.tsx")) is False

    def test_file_in_test_directory(self):
        """Should identify files in test directories as test files."""
        assert is_test_file(Path("/src/test/Sample.java")) is True
        assert is_test_file(Path("/src/tests/Sample.java")) is True
        assert is_test_file(Path("/src/__tests__/sample.js")) is True
        assert is_test_file(Path("/src/spec/sample.js")) is True
        assert is_test_file(Path("/src/specs/sample.js")) is True

    def test_file_in_test_directory_case_insensitive(self):
        """Should identify files in test directories (case-insensitive)."""
        assert is_test_file(Path("/src/Test/Sample.java")) is True
        assert is_test_file(Path("/src/Tests/Sample.java")) is True
        assert is_test_file(Path("/src/TESTS/Sample.java")) is True

    def test_java_maven_gradle_structure(self):
        """Should handle standard Java Maven/Gradle src/test directory structure."""
        # Files in src/test should be excluded regardless of filename
        assert is_test_file(Path("src/test/java/com/example/CalculatorTest.java")) is True
        assert is_test_file(Path("src/test/java/com/example/Helper.java")) is True
        assert is_test_file(Path("src/test/java/Utils.java")) is True

        # Files in src/main should be included
        assert is_test_file(Path("src/main/java/com/example/Calculator.java")) is False
        assert is_test_file(Path("src/main/java/com/example/Service.java")) is False

    def test_python_test_file_with_test_prefix(self):
        """Should identify Python files starting with test_ as test files."""
        assert is_test_file(Path("/src/test_utils.py")) is True
        assert is_test_file(Path("/src/test_calculator.py")) is True
        assert is_test_file(Path("/src/test_main.py")) is True

    def test_python_test_file_with_test_suffix(self):
        """Should identify Python files ending with _test.py or _tests.py as test files."""
        assert is_test_file(Path("/src/calculator_test.py")) is True
        assert is_test_file(Path("/src/utils_test.py")) is True
        assert is_test_file(Path("/src/calculator_tests.py")) is True
        assert is_test_file(Path("/src/utils_tests.py")) is True

    def test_python_non_test_file(self):
        """Should not identify regular Python files as test files."""
        assert is_test_file(Path("/src/calculator.py")) is False
        assert is_test_file(Path("/src/utils.py")) is False
        assert is_test_file(Path("/src/main.py")) is False
        # testing_utils.py should NOT be a test file (it's a utility module)
        assert is_test_file(Path("/src/testing_utils.py")) is False

    def test_python_file_in_test_directory(self):
        """Should identify Python files in test directories as test files."""
        assert is_test_file(Path("/src/tests/test_calculator.py")) is True
        assert is_test_file(Path("/src/tests/conftest.py")) is True
        assert is_test_file(Path("/src/tests/helper.py")) is True


class TestExcludeTestFiles:
    """Tests for ensuring test files are excluded from repository scanning."""

    def test_exclude_java_test_files(self, tmp_path: Path):
        """Should exclude Java test files from scanning."""
        repo_dir = tmp_path / "java_repo"
        src_dir = repo_dir / "src" / "main" / "java"
        test_dir = repo_dir / "src" / "test" / "java"
        src_dir.mkdir(parents=True)
        test_dir.mkdir(parents=True)

        # Create a regular source file
        source_content = """
public class Calculator {
    public int add(int a, int b) {
        return a + b;
    }
}
"""
        (src_dir / "Calculator.java").write_text(source_content)

        # Create test files that should be excluded
        test_content = """
public class CalculatorTest {
    public void testAdd() {
        Calculator calc = new Calculator();
        assert calc.add(2, 3) == 5;
    }
}
"""
        (test_dir / "CalculatorTest.java").write_text(test_content)
        (src_dir / "UtilsTest.java").write_text(test_content)

        # Scan the repository
        repo_name, functions = scan_single_repository(str(repo_dir))

        # Should find functions from Calculator.java but not from test files
        assert repo_name is not None
        func_files = [f.file_path for f in functions]

        # Should have function from Calculator.java
        assert any("Calculator.java" in f and "Test" not in f for f in func_files)

        # Should NOT have functions from test files
        assert not any("CalculatorTest.java" in f for f in func_files)
        assert not any("UtilsTest.java" in f for f in func_files)

    def test_exclude_javascript_test_files(self, tmp_path: Path):
        """Should exclude JavaScript test files from scanning."""
        repo_dir = tmp_path / "js_repo"
        src_dir = repo_dir / "src"
        src_dir.mkdir(parents=True)

        # Create a regular source file
        source_content = """
function calculate(a, b) {
    return a + b;
}

const multiply = (a, b) => {
    return a * b;
};
"""
        (src_dir / "utils.js").write_text(source_content)

        # Create test files that should be excluded
        test_content = """
function testCalculate() {
    expect(calculate(2, 3)).toBe(5);
}

const testMultiply = () => {
    expect(multiply(2, 3)).toBe(6);
};
"""
        (src_dir / "utils.test.js").write_text(test_content)
        (src_dir / "utils.spec.js").write_text(test_content)

        # Scan the repository
        repo_name, functions = scan_single_repository(str(repo_dir))

        # Should find functions from utils.js but not from test files
        assert repo_name is not None
        func_files = [f.file_path for f in functions]

        # Should have functions from utils.js
        assert any("utils.js" in f and ".test." not in f and ".spec." not in f for f in func_files)

        # Should NOT have functions from test files
        assert not any("utils.test.js" in f for f in func_files)
        assert not any("utils.spec.js" in f for f in func_files)

    def test_exclude_files_in_tests_directory(self, tmp_path: Path):
        """Should exclude all files in test directories."""
        repo_dir = tmp_path / "mixed_repo"
        src_dir = repo_dir / "src"
        test_dir = repo_dir / "tests"
        src_dir.mkdir(parents=True)
        test_dir.mkdir(parents=True)

        # Create source files
        (src_dir / "app.js").write_text("function app() { return 'app'; }")

        # Create files in test directory (should be excluded even without test naming)
        (test_dir / "helper.js").write_text("function helper() { return 'test helper'; }")
        (test_dir / "Helper.java").write_text("public class Helper { public void help() {} }")

        # Scan the repository
        repo_name, functions = scan_single_repository(str(repo_dir))

        assert repo_name is not None
        func_files = [f.file_path for f in functions]

        # Should have function from app.js
        assert any("app.js" in f for f in func_files)

        # Should NOT have functions from files in tests directory
        assert not any("tests" in Path(f).parts for f in func_files)

    def test_exclude_python_test_files(self, tmp_path: Path):
        """Should exclude Python test files from scanning."""
        repo_dir = tmp_path / "python_repo"
        src_dir = repo_dir / "src"
        src_dir.mkdir(parents=True)

        # Create a regular source file
        source_content = """
def calculate(a, b):
    return a + b

def multiply(a, b):
    return a * b
"""
        (src_dir / "utils.py").write_text(source_content)

        # Create test files that should be excluded
        test_content = """
def test_calculate():
    assert calculate(2, 3) == 5

def test_multiply():
    assert multiply(2, 3) == 6
"""
        (src_dir / "test_utils.py").write_text(test_content)
        (src_dir / "utils_test.py").write_text(test_content)

        # Scan the repository
        repo_name, functions = scan_single_repository(str(repo_dir))

        # Should find functions from utils.py but not from test files
        assert repo_name is not None
        func_files = [f.file_path for f in functions]

        # Should have functions from utils.py
        assert any("utils.py" in f and "test_" not in f and "_test.py" not in f for f in func_files)

        # Should NOT have functions from test files
        assert not any("test_utils.py" in f for f in func_files)
        assert not any("utils_test.py" in f for f in func_files)
